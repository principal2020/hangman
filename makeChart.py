# -*- coding: utf-8 -*-

import openpyxl as opxl

#wb = opxl.load_workbook('AthomeList.xlsx')
#sheet=wb.active
#typ=[]
#count=[]
#for i in range(1,75):
#    typ.append(sheet['c' + str(i)].value)
#    count.append(sheet['n' + str(i)].value)
#wb.close()

#for i in range(1,75):
#    print(count[i])

#wb =opxl.Workbook()
#sheet = wb.active
#for i in range(0,74):
#    sheet['a' + str(i+2)] = typ[i]
#    sheet['b' + str(i+2)] = count[i]
#ref = opxl.chart.Reference(sheet,min_col=1,min_row=1,max_col=1,max_row=10)
#ser = opxl.chart.Series(ref,title='Athome Analized Chart')
#chart=opxl.chart.BarChart()
#chart.append(ser)
#chart.y=50
#chart.x=100
#chart.w=300
#chart.h=200
#sheet.add_chart(chart)
#wb.save('Athome.xlsx')


wb = opxl.load_workbook('AthomeList.xlsx')
sheet=wb.active
count_type ={}
count_data = {}
for i in range(2,75):
    typ=sheet['c' + str(i)].value
    place=sheet['e' + str(i)].value
    count=sheet['n' + str(i)].value
    count_data.setdefault(typ,{'count':0})
    count_type.setdefault(typ,{'count':0})
#    count_data[typ].setdefault(typ,{'count':0})
    count_type[typ]['count'] += 1
    count_data[typ]['count'] += int(count)    
wb.close()
print(count_data['売地']['count'])
print(count_type['売地']['count'])
print(count_data['中古売戸建住宅']['count'])
print(count_type['中古売戸建住宅']['count'])
print(count_data['中古売マンション']['count'])
print(count_type['中古売マンション']['count'])
print(count_data['新築売戸建住宅']['count'])
print(count_type['新築売戸建住宅']['count'])